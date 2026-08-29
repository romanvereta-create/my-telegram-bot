import subprocess
import sys
import os

try:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-r", "requirements.txt"])
except:
    pass

import datetime
import shutil
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, BotCommand
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

def get_moscow_time():
    return datetime.datetime.now() + datetime.timedelta(hours=3)

class PDF(FPDF):
    def __init__(self):
        super().__init__('P', 'mm', (80, 250))
        self.add_font('DejaVu', '', 'DejaVuSansCondensed.ttf', uni=True)
        self.add_font('DejaVu', 'B', 'DejaVuSansCondensed-Bold.ttf', uni=True)

def load_settings():
    default_settings = {
        "company_name": "ИП Иванов И.И.",
        "inn": "1234567890",
        "ogrnip": "312345678901234",
        "address": "г. Москва, Волгоградский пр-т, 42, к. 9",
        "phone": "+7 (495) 123-45-67",
        "service_name": "Услуга",
        "tax_system": "ПСН",
        "email_sender": "ivanov@gmail.com",
        "thanks_text": "СПАСИБО ЗА ОПЛАТУ!",
        "website": "https://ваш-сайт.рф",
        "bank_name": "Т-Банк (АО)",
        "bik": "044525974",
        "account_number": "40802810500000012345",
        "corr_account": "30101810145250000974",
        "recipient": "ИП Иванов И.И."
    }

    settings = default_settings.copy()

    try:
        if os.path.exists('settings.xlsx'):
            wb = openpyxl.load_workbook('settings.xlsx')
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    key = str(row[0]).strip()
                    value = str(row[1]).strip()
                    settings[key] = value
            print("✅ Настройки загружены из settings.xlsx")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Настройки"
            headers = ["Параметр", "Значение"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            row = 2
            for key, value in default_settings.items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=value)
                row += 1
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 45
            wb.save('settings.xlsx')
            print("📄 Создан шаблон settings.xlsx")
    except Exception as e:
        print(f"⚠️ Ошибка загрузки settings.xlsx: {e}")

    return settings

def init_excel():
    if not os.path.exists('book.xlsx'):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Книга учёта"
        headers = ["№ п/п", "Дата и время", "№ Квитанции", "Клиент", "доходы, руб", "Статус"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        wb.save('book.xlsx')
        print("📄 Создан файл book.xlsx")

def add_to_excel(client_name, amount, receipt_number):
    try:
        init_excel()
        wb = openpyxl.load_workbook('book.xlsx')
        ws = wb.active
        now = get_moscow_time()
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=now.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row_num, column=3, value=receipt_number)
        ws.cell(row=row_num, column=4, value=client_name)
        ws.cell(row=row_num, column=5, value=amount)
        ws.cell(row=row_num, column=6, value="Оплачено")
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center')
        wb.save('book.xlsx')
        print(f"📝 Добавлена запись: {client_name} | {amount} руб. | №{receipt_number}")
    except PermissionError:
        print("⚠️ Ошибка: файл book.xlsx открыт в Excel!")
        if os.path.exists('book.xlsx'):
            shutil.copy('book.xlsx', 'book_backup.xlsx')
            print("📋 Создана резервная копия book_backup.xlsx")
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Книга учёта"
            headers = ["№ п/п", "Дата и время", "№ Квитанции", "Клиент", "доходы, руб", "Статус"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            now = get_moscow_time()
            row_num = 2
            ws.cell(row=row_num, column=1, value=1)
            ws.cell(row=row_num, column=2, value=now.strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row_num, column=3, value=receipt_number)
            ws.cell(row=row_num, column=4, value=client_name)
            ws.cell(row=row_num, column=5, value=amount)
            ws.cell(row=row_num, column=6, value="Оплачено")
            wb.save('book_new.xlsx')
            print("📋 Создан новый файл book_new.xlsx")
        except:
            print("❌ Не удалось сохранить Excel-файл.")
    except Exception as e:
        print(f"⚠️ Ошибка записи в Excel: {e}")

def load_clients():
    clients = {}
    try:
        if os.path.exists('clients.xlsx'):
            wb = openpyxl.load_workbook('clients.xlsx')
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    name = str(row[0]).strip()
                    try:
                        tid = int(row[1])
                        clients[name] = tid
                    except:
                        pass
            print(f"📂 Загружено {len(clients)} клиентов из clients.xlsx")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Клиенты"
            headers = ["Имя", "Telegram ID"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            wb.save('clients.xlsx')
            print("📄 Создан шаблон clients.xlsx")
    except Exception as e:
        print(f"⚠️ Ошибка загрузки clients.xlsx: {e}")
    return clients

def save_clients(clients):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    headers = ["Имя", "Telegram ID"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    
    row = 2
    for name, tid in clients.items():
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=tid)
        row += 1
    
    wb.save('clients.xlsx')

def generate_pdf(settings, client_name, amount, user_id):
    pdf = PDF()
    pdf.add_page()
    pdf.set_left_margin(5)
    pdf.set_right_margin(5)

    if os.path.exists('logo.png'):
        try:
            pdf.image('logo.png', x=30, y=5, w=20)
            pdf.ln(14)
        except:
            pdf.ln(5)
    else:
        pdf.ln(5)

    pdf.set_font('DejaVu', 'B', 12)
    pdf.cell(0, 5, txt="═" * 30, ln=True, align='C')
    pdf.set_font('DejaVu', 'B', 14)
    pdf.cell(0, 7, txt="КАССОВЫЙ ЧЕК", ln=True, align='C')
    pdf.set_font('DejaVu', 'B', 12)
    pdf.cell(0, 5, txt="═" * 30, ln=True, align='C')
    pdf.ln(2)

    now = get_moscow_time()
    receipt_number = f"{now.strftime('%d%m')}-{now.strftime('%H%M%S')}"
    pdf.set_font('DejaVu', '', 9)
    pdf.cell(0, 5, txt=f"Чек №: {receipt_number}", ln=True, align='C')
    pdf.cell(0, 5, txt=f"{now.strftime('%d.%m.%Y %H:%M')}", ln=True, align='C')
    pdf.ln(2)

    pdf.set_font('DejaVu', '', 8)
    pdf.cell(0, 4, txt="─" * 30, ln=True, align='C')
    pdf.ln(2)

    pdf.set_font('DejaVu', 'B', 10)
    pdf.cell(0, 5, txt=settings["company_name"], ln=True, align='C')
    pdf.set_font('DejaVu', '', 8)
    pdf.cell(0, 4, txt=f"ИНН: {settings['inn']}", ln=True, align='C')
    if settings.get('ogrnip'):
        pdf.cell(0, 4, txt=f"ОГРНИП: {settings['ogrnip']}", ln=True, align='C')
    pdf.cell(0, 4, txt=settings["address"], ln=True, align='C')
    if settings.get('phone'):
        pdf.cell(0, 4, txt=f"Тел.: {settings['phone']}", ln=True, align='C')
    pdf.ln(2)

    pdf.set_font('DejaVu', '', 8)
    pdf.cell(0, 4, txt="─" * 30, ln=True, align='C')
    pdf.ln(2)

    pdf.set_font('DejaVu', 'B', 8)
    pdf.cell(40, 5, "Наименование", border=1, align='C')
    pdf.cell(20, 5, "Цена", border=1, align='C')
    pdf.cell(10, 5, "Кол", border=1, align='C')
    pdf.ln()

    pdf.set_font('DejaVu', '', 8)
    pdf.cell(40, 5, settings['service_name'], border=1, align='L')
    pdf.cell(20, 5, f"{amount:.2f}", border=1, align='R')
    pdf.cell(10, 5, "1", border=1, align='C')
    pdf.ln()
    pdf.ln(2)

    pdf.set_font('DejaVu', '', 9)
    pdf.cell(50, 5, txt="СУММА БЕЗ НДС", border=0)
    pdf.cell(20, 5, txt=f"{amount:.2f}", border=0, align='R')
    pdf.ln()
    pdf.set_font('DejaVu', 'B', 10)
    pdf.cell(50, 6, txt="ИТОГО:", border=0)
    pdf.cell(20, 6, txt=f"{amount:.2f}", border=0, align='R')
    pdf.ln()
    pdf.set_font('DejaVu', '', 8)
    pdf.cell(50, 4, txt="Безналичными", border=0)
    pdf.cell(20, 4, txt=f"{amount:.2f}", border=0, align='R')
    pdf.ln(3)

    pdf.set_font('DejaVu', '', 8)
    pdf.cell(0, 4, txt="─" * 30, ln=True, align='C')
    pdf.ln(2)

    pdf.set_font('DejaVu', '', 8)
    pdf.cell(0, 4, txt=f"Система: {settings['tax_system']}", ln=True, align='C')
    pdf.ln(2)

    pdf.set_font('DejaVu', '', 8)
    pdf.cell(0, 4, txt="─" * 30, ln=True, align='C')
    pdf.ln(2)

    pdf.set_font('DejaVu', 'B', 8)
    pdf.cell(0, 4, txt="РЕКВИЗИТЫ:", ln=True, align='C')
    pdf.set_font('DejaVu', '', 7)
    if settings.get('bank_name'):
        pdf.cell(0, 3, txt=f"{settings['bank_name']}", ln=True, align='C')
    if settings.get('bik'):
        pdf.cell(0, 3, txt=f"БИК: {settings['bik']}", ln=True, align='C')
    if settings.get('account_number'):
        pdf.cell(0, 3, txt=f"Сч: {settings['account_number']}", ln=True, align='C')
    if settings.get('recipient'):
        pdf.cell(0, 3, txt=f"Получатель: {settings['recipient']}", ln=True, align='C')
    pdf.ln(2)

    pdf.set_font('DejaVu', '', 7)
    pdf.cell(0, 4, txt="─" * 30, ln=True, align='C')
    pdf.ln(2)

    pdf.set_font('DejaVu', '', 7)
    pdf.cell(0, 4, txt=f"Email: {settings['email_sender']}", ln=True, align='C')

    if settings.get('website'):
        pdf.cell(0, 4, txt=f"{settings['website']}", ln=True, align='C')

    if os.path.exists('signature.png'):
        try:
            pdf.image('signature.png', x=15, y=pdf.get_y() + 2, w=50)
            pdf.set_y(pdf.get_y() + 12)
        except:
            pdf.set_font('DejaVu', '', 7)
            pdf.cell(0, 4, txt="____________________", ln=True, align='C')
            pdf.cell(0, 4, txt="(подпись ИП)", ln=True, align='C')
    else:
        pdf.set_font('DejaVu', '', 7)
        pdf.cell(0, 4, txt="____________________", ln=True, align='C')
        pdf.cell(0, 4, txt="(подпись ИП)", ln=True, align='C')

    if os.path.exists('qrcode.png'):
        try:
            pdf.image('qrcode.png', x=30, y=pdf.get_y() + 22, w=20)
        except:
            pass

    filename = f"check_{user_id}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    pdf.output(filename)
    return filename, receipt_number

def get_main_keyboard():
    buttons = [
        [InlineKeyboardButton("📤 Выдать чек", callback_data="issue_cheque")],
        [InlineKeyboardButton("👥 Клиенты", callback_data="manage_clients")],
        [InlineKeyboardButton("📥 Скачать книгу", callback_data="download_book")],
        [InlineKeyboardButton("⚙️ Настройки", callback_data="settings")]
    ]
    return InlineKeyboardMarkup(buttons)

def get_clients_keyboard():
    clients = load_clients()
    buttons = []
    for name in clients.keys():
        buttons.append([InlineKeyboardButton(f"👤 {name}", callback_data=f"client_{name}")])
    buttons.append([InlineKeyboardButton("➕ Добавить", callback_data="add_client")])
    buttons.append([InlineKeyboardButton("🔙 Назад", callback_data="back_to_main")])
    return InlineKeyboardMarkup(buttons)

def get_client_actions_keyboard(client_name):
    buttons = [
        [InlineKeyboardButton("✅ Выдать чек", callback_data=f"issue_for_{client_name}")],
        [InlineKeyboardButton("🗑 Удалить", callback_data=f"delete_client_{client_name}")],
        [InlineKeyboardButton("🔙 Назад", callback_data="back_to_clients")]
    ]
    return InlineKeyboardMarkup(buttons)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = str(update.effective_user.id)
    first_name = update.effective_user.first_name or "Пользователь"
    
    clients = load_clients()
    exists = False
    for name, tid in clients.items():
        if str(tid) == user_id:
            exists = True
            break
    if not exists:
        clients[first_name] = int(user_id)
        save_clients(clients)
        await update.message.reply_text(f"✅ {first_name}, ты зарегистрирован!")
    
    commands = [
        BotCommand("start", "🏠 Главное меню"),
        BotCommand("clients", "👥 Клиенты"),
        BotCommand("book", "📥 Скачать книгу"),
        BotCommand("settings", "⚙️ Настройки")
    ]
    await context.bot.set_my_commands(commands)
    
    await update.message.reply_text(
        "🏠 *Главное меню*\n\nВыберите действие:",
        reply_markup=get_main_keyboard(),
        parse_mode='Markdown'
    )

async def main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        "🏠 *Главное меню*\n\nВыберите действие:",
        reply_markup=get_main_keyboard(),
        parse_mode='Markdown'
    )

async def manage_clients(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    clients = load_clients()
    if not clients:
        await query.edit_message_text(
            "👥 *Нет клиентов*\n\nДобавьте первого клиента:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("➕ Добавить", callback_data="add_client")],
                [InlineKeyboardButton("🔙 Назад", callback_data="back_to_main")]
            ]),
            parse_mode='Markdown'
        )
        return
    
    text = "👥 *Список клиентов*\n\n"
    for name in clients.keys():
        text += f"• {name}\n"
    text += f"\nВсего: {len(clients)}"
    
    await query.edit_message_text(
        text,
        reply_markup=get_clients_keyboard(),
        parse_mode='Markdown'
    )

async def client_actions(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    client_name = query.data.replace("client_", "")
    context.user_data['selected_client'] = client_name
    await query.edit_message_text(
        f"👤 *{client_name}*\n\nВыберите действие:",
        reply_markup=get_client_actions_keyboard(client_name),
        parse_mode='Markdown'
    )

async def delete_client(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    client_name = query.data.replace("delete_client_", "")
    clients = load_clients()
    if client_name in clients:
        del clients[client_name]
        save_clients(clients)
        await query.edit_message_text(
            f"✅ *{client_name}* удалён!",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 К списку", callback_data="back_to_clients")]
            ]),
            parse_mode='Markdown'
        )
    else:
        await query.edit_message_text("❌ Клиент не найден")

async def add_client(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        "✏️ *Добавление клиента*\n\nВведите имя и ID через запятую:\n`Имя, 123456789`\n\nID можно узнать, попросив написать /start",
        parse_mode='Markdown'
    )
    context.user_data['waiting_for_client'] = True

async def handle_client_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get('waiting_for_client'):
        return
    try:
        parts = update.message.text.split(',')
        if len(parts) < 2:
            await update.message.reply_text("❌ Формат: `Имя, 123456789`")
            return
        name = parts[0].strip()
        tid = int(parts[1].strip())
        clients = load_clients()
        clients[name] = tid
        save_clients(clients)
        context.user_data.pop('waiting_for_client', None)
        await update.message.reply_text(f"✅ *{name}* добавлен!", parse_mode='Markdown')
        await main_menu(update, context)
    except ValueError:
        await update.message.reply_text("❌ ID должен быть числом")
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {e}")

async def issue_cheque(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    clients = load_clients()
    if not clients:
        await query.edit_message_text(
            "❌ *Нет клиентов*\n\nСначала добавьте клиентов.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Назад", callback_data="back_to_main")]
            ]),
            parse_mode='Markdown'
        )
        return
    buttons = []
    for name in clients.keys():
        buttons.append([InlineKeyboardButton(f"👤 {name}", callback_data=f"issue_for_{name}")])
    buttons.append([InlineKeyboardButton("✏️ Вручную", callback_data="issue_manual")])
    buttons.append([InlineKeyboardButton("🔙 Назад", callback_data="back_to_main")])
    await query.edit_message_text(
        "👤 *Выберите клиента:*",
        reply_markup=InlineKeyboardMarkup(buttons),
        parse_mode='Markdown'
    )

async def issue_manual(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("✏️ *Введите имя клиента:*", parse_mode='Markdown')
    context.user_data['waiting_for_manual_client'] = True

async def ask_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    client_name = query.data.replace("issue_for_", "")
    context.user_data['client_for_cheque'] = client_name
    await query.edit_message_text(
        f"💳 *Чек для {client_name}*\n\nВведите сумму:",
        parse_mode='Markdown'
    )
    context.user_data['waiting_for_amount'] = True

async def handle_amount_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get('waiting_for_amount') and not context.user_data.get('waiting_for_manual_amount'):
        return
    try:
        amount = float(update.message.text.replace(',', '.'))
        if amount <= 0:
            await update.message.reply_text("❌ Сумма должна быть > 0")
            return
        client_name = context.user_data.get('client_for_cheque')
        if not client_name:
            await update.message.reply_text("❌ Ошибка: клиент не выбран")
            return
        clients = load_clients()
        client_id = clients.get(client_name)
        settings = load_settings()
        pdf_file, receipt_number = generate_pdf(settings, client_name, amount, update.effective_user.id)
        add_to_excel(client_name, amount, receipt_number)
        with open(pdf_file, 'rb') as f:
            await update.message.reply_document(document=f, filename=pdf_file)
        await update.message.reply_text(
            f"✅ Чек для *{client_name}* на *{amount:.2f} руб.*\n№ {receipt_number}",
            parse_mode='Markdown'
        )
        if client_id:
            try:
                with open(pdf_file, 'rb') as f:
                    await context.bot.send_document(chat_id=client_id, document=f, filename=pdf_file)
                await update.message.reply_text(f"📤 Отправлен *{client_name}*", parse_mode='Markdown')
            except:
                await update.message.reply_text("⚠️ Не удалось отправить клиенту")
        else:
            await update.message.reply_text("⚠️ Клиент не найден, чек только вам")
        try:
            os.remove(pdf_file)
        except:
            pass
        context.user_data.pop('waiting_for_amount', None)
        context.user_data.pop('waiting_for_manual_amount', None)
        context.user_data.pop('client_for_cheque', None)
        await update.message.reply_text(
            "🏠 *Главное меню*",
            reply_markup=get_main_keyboard(),
            parse_mode='Markdown'
        )
    except ValueError:
        await update.message.reply_text("❌ Введите число (например: 1500)")

async def download_book(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not os.path.exists('book.xlsx'):
        await query.edit_message_text("❌ Книга учёта ещё не создана")
        return
    try:
        with open('book.xlsx', 'rb') as f:
            await query.message.reply_document(
                document=f,
                filename=f"книга_учёта_{datetime.datetime.now().strftime('%d.%m.%Y')}.xlsx",
                caption="📋 Книга учёта"
            )
        await query.edit_message_text(
            "✅ Книга учёта отправлена!",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 Назад", callback_data="back_to_main")]
            ])
        )
    except Exception as e:
        await query.edit_message_text(f"❌ Ошибка: {e}")

async def settings_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        "⚙️ *Настройки*\n\nРеквизиты редактируются в `settings.xlsx` на сервере.",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🔙 Назад", callback_data="back_to_main")]
        ]),
        parse_mode='Markdown'
    )

async def back_to_clients(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await manage_clients(update, context)

async def back_to_main(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await main_menu(update, context)

async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data
    
    if data == "back_to_main":
        await main_menu(update, context)
        return
    if data == "back_to_clients":
        await manage_clients(update, context)
        return
    if data == "manage_clients":
        await manage_clients(update, context)
        return
    if data == "add_client":
        await add_client(update, context)
        return
    if data == "issue_cheque":
        await issue_cheque(update, context)
        return
    if data == "issue_manual":
        await issue_manual(update, context)
        return
    if data == "download_book":
        await download_book(update, context)
        return
    if data == "settings":
        await settings_menu(update, context)
        return
    if data.startswith("client_"):
        await client_actions(update, context)
        return
    if data.startswith("delete_client_"):
        await delete_client(update, context)
        return
    if data.startswith("issue_for_"):
        await ask_amount(update, context)
        return

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get('waiting_for_client'):
        await handle_client_input(update, context)
        return
    if context.user_data.get('waiting_for_manual_client'):
        client_name = update.message.text.strip()
        if client_name:
            context.user_data['client_for_cheque'] = client_name
            context.user_data.pop('waiting_for_manual_client', None)
            context.user_data['waiting_for_manual_amount'] = True
            await update.message.reply_text(
                f"💳 *Чек для {client_name}*\n\nВведите сумму:",
                parse_mode='Markdown'
            )
        return
    if context.user_data.get('waiting_for_amount') or context.user_data.get('waiting_for_manual_amount'):
        await handle_amount_input(update, context)
        return

def main():
    TOKEN = os.getenv('BOT_TOKEN')
    if not TOKEN:
        print("❌ BOT_TOKEN не найден!")
        return
    init_excel()
    load_settings()
    load_clients()
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("clients", manage_clients))
    app.add_handler(CommandHandler("book", download_book))
    app.add_handler(CommandHandler("settings", settings_menu))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    print("✅ БОТ ЗАПУЩЕН!")
    app.run_polling()

if __name__ == "__main__":
    main()
